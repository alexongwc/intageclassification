import streamlit as st
import os
import json
import re
import csv
from openai import OpenAI
from unstructured.partition.docx import partition_docx
from unstructured.documents.elements import Text, Title, NarrativeText
import openpyxl
import pandas as pd

VLLM_API_URL = "http://localhost:8000/v1"
MODEL_NAME = "/home/work/IntageAudit/classification/model/Qwen2.5-32B-Instruct"
API_KEY = "dummy-key"
CHECKLIST_PATH = "checklist_items.json"

client = OpenAI(base_url=VLLM_API_URL, api_key=API_KEY)

def extract_and_format_transcript(docx_path):
    elements = partition_docx(docx_path)
    raw_text = "\n".join([el.text for el in elements if isinstance(el, (Text, Title, NarrativeText))])
    parts = re.split(r'(MS:|A:)', raw_text)
    formatted_lines = []
    current_speaker = None
    for i, part in enumerate(parts):
        part = part.strip()
        if not part:
            continue
        if part in ['MS:', 'A:']:
            current_speaker = part[:-1]
        elif current_speaker and part:
            text = part.strip()
            if text:
                formatted_lines.append(f"[{current_speaker}]: {text}")
    return "\n".join(formatted_lines)

def extract_and_format_excel(excel_path):
    try:
        df = pd.read_excel(excel_path, sheet_name=None)
        formatted_lines = []
        
        for sheet_name, sheet_df in df.items():
            if sheet_df.empty:
                continue
                
            formatted_lines.append(f"=== SHEET: {sheet_name} ===")
            
            # Check for timestamp columns
            timestamp_cols = [col for col in sheet_df.columns if any(keyword in col.lower() for keyword in ['time', 'timestamp', 'date', 'when', 'at'])]
            
            for index, row in sheet_df.iterrows():
                row_parts = []
                timestamp_info = ""
                
                # Extract timestamp information first if available
                if timestamp_cols:
                    timestamps = []
                    for ts_col in timestamp_cols:
                        if pd.notna(row[ts_col]):
                            timestamps.append(f"{ts_col}: {str(row[ts_col])}")
                    if timestamps:
                        timestamp_info = f"[TIMESTAMP: {' | '.join(timestamps)}] "
                
                # Format all other columns
                for col, val in row.items():
                    if pd.notna(val):
                        row_parts.append(f"{col}: {str(val)}")
                
                if row_parts:
                    formatted_lines.append(f"Row {index + 1}: {timestamp_info}{' | '.join(row_parts)}")
        
        return "\n".join(formatted_lines)
    except Exception as e:
        st.error(f"Error processing Excel file: {str(e)}")
        return ""

def extract_and_format_csv(csv_path):
    try:
        df = pd.read_csv(csv_path)
        formatted_lines = []
        
        # Check for timestamp columns
        timestamp_cols = [col for col in df.columns if any(keyword in col.lower() for keyword in ['time', 'timestamp', 'date', 'when', 'at'])]
        
        for index, row in df.iterrows():
            row_parts = []
            timestamp_info = ""
            
            # Extract timestamp information first if available
            if timestamp_cols:
                timestamps = []
                for ts_col in timestamp_cols:
                    if pd.notna(row[ts_col]):
                        timestamps.append(f"{ts_col}: {str(row[ts_col])}")
                if timestamps:
                    timestamp_info = f"[TIMESTAMP: {' | '.join(timestamps)}] "
            
            # Format all other columns
            for col, val in row.items():
                if pd.notna(val):
                    row_parts.append(f"{col}: {str(val)}")
            
            if row_parts:
                formatted_lines.append(f"Row {index + 1}: {timestamp_info}{' | '.join(row_parts)}")
        
        return "\n".join(formatted_lines)
    except Exception as e:
        st.error(f"Error processing CSV file: {str(e)}")
        return ""

def extract_and_format_json(json_path):
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        def format_json_data(obj, prefix=""):
            lines = []
            if isinstance(obj, dict):
                for key, value in obj.items():
                    if isinstance(value, (dict, list)):
                        lines.append(f"{prefix}{key}:")
                        lines.extend(format_json_data(value, prefix + "  "))
                    else:
                        lines.append(f"{prefix}{key}: {value}")
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    if isinstance(item, (dict, list)):
                        lines.append(f"{prefix}[{i}]:")
                        lines.extend(format_json_data(item, prefix + "  "))
                    else:
                        lines.append(f"{prefix}[{i}]: {item}")
            else:
                lines.append(f"{prefix}{obj}")
            return lines
        
        return "\n".join(format_json_data(data))
    except Exception as e:
        st.error(f"Error processing JSON file: {str(e)}")
        return ""

def load_and_flatten_checklist(checklist_path):
    with open(checklist_path, 'r') as f:
        original_checklist = json.load(f)
    flat_items = []
    stack = [(item, "") for item in reversed(original_checklist)]
    while stack:
        current_item, parent_context = stack.pop()
        current_question = f"{parent_context} > {current_item['question_text']}" if parent_context else current_item['question_text']
        if current_item.get('sub_items'):
            for sub_item in reversed(current_item['sub_items']):
                stack.append((sub_item, current_question))
        else:
            flat_items.append({
                'id': current_item['id'],
                'full_question': current_question,
                'original_item': current_item
            })
    return original_checklist, flat_items

def check_single_checklist_item(item, transcript):
    estimated_tokens = len(transcript) // 4
    if estimated_tokens < 14000:
        prompt = f"""You are a careful analyst reviewing a conversation transcript.

CONVERSATION TRANSCRIPT:
{transcript}

YOUR TASK:
Check if the following checklist item is addressed ANYWHERE in the above conversation.

CHECKLIST ITEM TO VERIFY:
ID: {item['id']}
Question: {item['full_question']}

IMPORTANT INSTRUCTIONS:
1. Read the ENTIRE conversation carefully
2. The conversation is between different speakers marked with [SPEAKER_NAME]: 
3. Look for ANY part of the conversation that answers this specific question
4. Only mark as "answered" if you find DIRECT evidence
5. Extract the EXACT quotes from the conversation as evidence
6. MANDATORY: Every evidence quote MUST start with the speaker label [Mysteryshopper]: or [Insuranceagent]:
7. If you see timestamp information in the format [TIMESTAMP: ...], include it with the evidence
8. For structured data, look for Row numbers and timestamp markers to provide context

Critical: Output ONLY valid JSON, nothing else.

OUTPUT FORMAT:
{{
  "item_id": "{item['id']}",
  "is_answered": true,
  "evidence": ["[TIMESTAMP: time_info] exact quote from transcript", "Row X: [TIMESTAMP: time_info] data content"],
  "answer_summary": "Brief summary",
  "confidence": 0.8,
  "timestamps": ["time1", "time2"]
}}"""

        try:
            response = client.chat.completions.create(
                model=MODEL_NAME,
                messages=[
                    {"role": "system", "content": "You are a precise analyst. Output ONLY valid JSON, no additional text."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.0,
                max_tokens=1024,
                timeout=120
            )
            result_text = response.choices[0].message.content.strip()
            try:
                return json.loads(result_text)
            except json.JSONDecodeError:
                pass
            try:
                json_start = result_text.find('{')
                json_end = result_text.rfind('}')
                if json_start != -1 and json_end != -1:
                    potential_json = result_text[json_start : json_end + 1]
                    return json.loads(potential_json)
            except json.JSONDecodeError:
                pass
            print(f"[WARN] Failed to parse JSON for {item['id']}: {result_text[:200]}")
            return {
                "item_id": item['id'],
                "is_answered": False,
                "evidence": [],
                "answer_summary": "JSON parse error",
                "confidence": 0.0,
                "timestamps": []
            }
        except Exception as e:
            print(f"[ERROR] Failed for item {item['id']}: {str(e)}")
            return {
                "item_id": item['id'],
                "is_answered": False,
                "evidence": [],
                "answer_summary": f"Error: {str(e)}",
                "confidence": 0.0,
                "timestamps": []
            }
    else:
        lines = transcript.split('\n')
        chunks = []
        current_chunk = ""
        for line in lines:
            if len(current_chunk) + len(line) < 14000 * 3:
                current_chunk += line + '\n'
            else:
                chunks.append(current_chunk)
                current_chunk = line + '\n'
        chunks.append(current_chunk)
        last_result = None
        for chunk in chunks:
            result = check_single_checklist_item(item, chunk)
            if result.get('is_answered', False):
                return result
            last_result = result
        return last_result

def merge_results_into_checklist(original_checklist, results):
    results_map = {r['item_id']: r for r in results}
    def update_items(items):
        updated = []
        for item in items:
            item_copy = item.copy()
            if item['id'] in results_map:
                result = results_map[item['id']]
                item_copy['status'] = 'Checked' if result.get('is_answered') else 'Not Checked'
                item_copy['answer'] = result.get('answer_summary', 'Not found')
                item_copy['evidence'] = result.get('evidence', [])
                item_copy['confidence_score'] = result.get('confidence', 0.0)
                item_copy['timestamps'] = result.get('timestamps', [])
            else:
                item_copy['status'] = 'Not Checked'
                item_copy['answer'] = 'Not found in transcript'
                item_copy['evidence'] = []
                item_copy['confidence_score'] = 0.0
                item_copy['timestamps'] = []
            if item.get('sub_items'):
                item_copy['sub_items'] = update_items(item['sub_items'])
            updated.append(item_copy)
        return updated
    return update_items(original_checklist)

def create_excel_report(final_output, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist Results"
    headers = ["Item ID", "Question", "Status", "Answer", "Evidence", "Timestamps", "Confidence"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    flat_list = []
    items_with_level = [(item, 0) for item in final_output['checklist']]
    idx = 0
    while idx < len(items_with_level):
        item, level = items_with_level[idx]
        flat_list.append((item, level))
        if item.get('sub_items'):
            sub_items = [(sub_item, level + 1) for sub_item in item['sub_items']]
            items_with_level[idx+1:idx+1] = sub_items
        idx += 1
    current_row = 2
    for item, level in flat_list:
        question_text = "  " * level + item.get('question_text', '')
        ws.cell(row=current_row, column=1, value=item.get('id', ''))
        ws.cell(row=current_row, column=2, value=question_text)
        ws.cell(row=current_row, column=3, value=item.get('status', 'Not Checked'))
        ws.cell(row=current_row, column=4, value=item.get('answer', ''))
        ws.cell(row=current_row, column=5, value='\n'.join(item.get('evidence', [])))
        ws.cell(row=current_row, column=6, value='\n'.join(item.get('timestamps', [])))
        ws.cell(row=current_row, column=7, value=item.get('confidence_score', 0))
        current_row += 1
    wb.save(excel_path)

st.title("Scoring Application")
st.header("Upload File")
uploaded_file = st.file_uploader("Choose a file", type=['docx', 'xlsx', 'xls', 'csv', 'json'])
if uploaded_file:
    st.write(f"File uploaded: {uploaded_file.name}")
    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
    temp_file = f"temp_{uploaded_file.name}"
    with open(temp_file, 'wb') as f:
        f.write(uploaded_file.getbuffer())
if uploaded_file and st.button("Run Verification"):
    base_name = os.path.splitext(uploaded_file.name)[0]
    run_dir = f"runs/{base_name}"
    os.makedirs(run_dir, exist_ok=True)
    
    # Process file based on type
    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
    if file_extension == '.docx':
        transcript = extract_and_format_transcript(temp_file)
    elif file_extension in ['.xlsx', '.xls']:
        transcript = extract_and_format_excel(temp_file)
    elif file_extension == '.csv':
        transcript = extract_and_format_csv(temp_file)
    elif file_extension == '.json':
        transcript = extract_and_format_json(temp_file)
    else:
        st.error(f"Unsupported file type: {file_extension}")
        transcript = ""
    original_checklist, flat_items = load_and_flatten_checklist(CHECKLIST_PATH)
    results = []
    for item in flat_items:
        result = check_single_checklist_item(item, transcript)
        results.append(result)
    updated_checklist = merge_results_into_checklist(original_checklist, results)
    items_answered = sum(1 for r in results if r.get('is_answered'))
    avg_confidence = sum(r.get('confidence', 0) for r in results) / len(results) if results else 0
    final_output = {
        'metadata': {
            'transcript_file': uploaded_file.name,
            'checklist_file': CHECKLIST_PATH,
            'total_items': len(results),
            'items_answered': items_answered,
            'items_not_found': len(results) - items_answered,
            'average_confidence': avg_confidence
        },
        'checklist': updated_checklist
    }
    excel_path = f"{run_dir}/{base_name}.xlsx"
    create_excel_report(final_output, excel_path)
    st.session_state['excel_path'] = excel_path
if 'excel_path' in st.session_state:
    with open(st.session_state['excel_path'], 'rb') as f:
        st.download_button(
            label="Download Excel Report",
            data=f.read(),
            file_name=os.path.basename(st.session_state['excel_path']),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )