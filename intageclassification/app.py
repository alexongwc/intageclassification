import streamlit as st
import os
import json
import re
from openai import OpenAI
from unstructured.partition.docx import partition_docx
from unstructured.documents.elements import Text, Title, NarrativeText
import openpyxl

VLLM_API_URL = "http://localhost:8000/v1"
MODEL_NAME = "/home/work/MAS/models/qwen2.5-32b"
API_KEY = "dummy-key"
CHECKLIST_PATH = "checklist_items.json"

client = OpenAI(base_url=VLLM_API_URL, api_key=API_KEY)

def extract_and_format_transcript(docx_path):
    elements = partition_docx(docx_path)
    raw_text = "\n".join([el.text for el in elements if isinstance(el, (Text, Title, NarrativeText))])
    parts = re.split(r'(MS:|A:)', raw_text)
    formatted_lines = []
    current_speaker = None
    for i, part in enumerate(parts):
        part = part.strip()
        if not part:
            continue
        if part in ['MS:', 'A:']:
            current_speaker = part[:-1]
        elif current_speaker and part:
            text = part.strip()
            if text:
                formatted_lines.append(f"[{current_speaker}]: {text}")
    return "\n".join(formatted_lines)

def load_and_flatten_checklist(checklist_path):
    with open(checklist_path, 'r') as f:
        original_checklist = json.load(f)
    flat_items = []
    stack = [(item, "") for item in reversed(original_checklist)]
    while stack:
        current_item, parent_context = stack.pop()
        current_question = f"{parent_context} > {current_item['question_text']}" if parent_context else current_item['question_text']
        if current_item.get('sub_items'):
            for sub_item in reversed(current_item['sub_items']):
                stack.append((sub_item, current_question))
        else:
            flat_items.append({
                'id': current_item['id'],
                'full_question': current_question,
                'original_item': current_item
            })
    return original_checklist, flat_items

def check_single_checklist_item(item, transcript):
    estimated_tokens = len(transcript) // 4
    if estimated_tokens < 14000:
        prompt = f"""You are a careful analyst reviewing a conversation transcript.

CONVERSATION TRANSCRIPT:
{transcript}

YOUR TASK:
Check if the following checklist item is addressed ANYWHERE in the above conversation.

CHECKLIST ITEM TO VERIFY:
ID: {item['id']}
Question: {item['full_question']}

IMPORTANT INSTRUCTIONS:
1. Read the ENTIRE conversation carefully
2. The conversation is between different speakers marked with [SPEAKER_NAME]: 
3. Look for ANY part of the conversation that answers this specific question
4. Only mark as "answered" if you find DIRECT evidence
5. Extract the EXACT quotes from the conversation as evidence
6. MANDATORY: Every evidence quote MUST start with the speaker label [MS]: or [A]:

Critical: Output ONLY valid JSON, nothing else.

OUTPUT FORMAT:
{{
  "item_id": "{item['id']}",
  "is_answered": true,
  "evidence": ["exact quote from transcript"],
  "answer_summary": "Brief summary",
  "confidence": 0.8
}}"""

        try:
            response = client.chat.completions.create(
                model=MODEL_NAME,
                messages=[
                    {"role": "system", "content": "You are a precise analyst. Output ONLY valid JSON, no additional text."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.0,
                max_tokens=1024,
                timeout=120
            )
            result_text = response.choices[0].message.content.strip()
            try:
                return json.loads(result_text)
            except json.JSONDecodeError:
                pass
            try:
                json_start = result_text.find('{')
                json_end = result_text.rfind('}')
                if json_start != -1 and json_end != -1:
                    potential_json = result_text[json_start : json_end + 1]
                    return json.loads(potential_json)
            except json.JSONDecodeError:
                pass
            print(f"[WARN] Failed to parse JSON for {item['id']}: {result_text[:200]}")
            return {
                "item_id": item['id'],
                "is_answered": False,
                "evidence": [],
                "answer_summary": "JSON parse error",
                "confidence": 0.0
            }
        except Exception as e:
            print(f"[ERROR] Failed for item {item['id']}: {str(e)}")
            return {
                "item_id": item['id'],
                "is_answered": False,
                "evidence": [],
                "answer_summary": f"Error: {str(e)}",
                "confidence": 0.0
            }
    else:
        lines = transcript.split('\n')
        chunks = []
        current_chunk = ""
        for line in lines:
            if len(current_chunk) + len(line) < 14000 * 3:
                current_chunk += line + '\n'
            else:
                chunks.append(current_chunk)
                current_chunk = line + '\n'
        chunks.append(current_chunk)
        last_result = None
        for chunk in chunks:
            result = check_single_checklist_item(item, chunk)
            if result.get('is_answered', False):
                return result
            last_result = result
        return last_result

def merge_results_into_checklist(original_checklist, results):
    results_map = {r['item_id']: r for r in results}
    def update_items(items):
        updated = []
        for item in items:
            item_copy = item.copy()
            if item['id'] in results_map:
                result = results_map[item['id']]
                item_copy['status'] = 'Checked' if result.get('is_answered') else 'Not Checked'
                item_copy['answer'] = result.get('answer_summary', 'Not found')
                item_copy['evidence'] = result.get('evidence', [])
                item_copy['confidence_score'] = result.get('confidence', 0.0)
            else:
                item_copy['status'] = 'Not Checked'
                item_copy['answer'] = 'Not found in transcript'
                item_copy['evidence'] = []
                item_copy['confidence_score'] = 0.0
            if item.get('sub_items'):
                item_copy['sub_items'] = update_items(item['sub_items'])
            updated.append(item_copy)
        return updated
    return update_items(original_checklist)

def create_excel_report(final_output, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist Results"
    headers = ["Item ID", "Question", "Status", "Answer", "Evidence", "Confidence"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    flat_list = []
    items_with_level = [(item, 0) for item in final_output['checklist']]
    idx = 0
    while idx < len(items_with_level):
        item, level = items_with_level[idx]
        flat_list.append((item, level))
        if item.get('sub_items'):
            sub_items = [(sub_item, level + 1) for sub_item in item['sub_items']]
            items_with_level[idx+1:idx+1] = sub_items
        idx += 1
    current_row = 2
    for item, level in flat_list:
        question_text = "  " * level + item.get('question_text', '')
        ws.cell(row=current_row, column=1, value=item.get('id', ''))
        ws.cell(row=current_row, column=2, value=question_text)
        ws.cell(row=current_row, column=3, value=item.get('status', 'Not Checked'))
        ws.cell(row=current_row, column=4, value=item.get('answer', ''))
        ws.cell(row=current_row, column=5, value='\n'.join(item.get('evidence', [])))
        ws.cell(row=current_row, column=6, value=item.get('confidence_score', 0))
        current_row += 1
    wb.save(excel_path)

st.title("Scoring Application")
st.header("Upload Transcript")
uploaded_file = st.file_uploader("Choose a DOCX file", type=['docx'])
if uploaded_file:
    st.write(f"File uploaded: {uploaded_file.name}")
    temp_docx = f"temp_{uploaded_file.name}"
    with open(temp_docx, 'wb') as f:
        f.write(uploaded_file.getbuffer())
if uploaded_file and st.button("Run Verification"):
    base_name = os.path.splitext(uploaded_file.name)[0]
    run_dir = f"runs/{base_name}"
    os.makedirs(run_dir, exist_ok=True)
    transcript = extract_and_format_transcript(temp_docx)
    original_checklist, flat_items = load_and_flatten_checklist(CHECKLIST_PATH)
    results = []
    for item in flat_items:
        result = check_single_checklist_item(item, transcript)
        results.append(result)
    updated_checklist = merge_results_into_checklist(original_checklist, results)
    items_answered = sum(1 for r in results if r.get('is_answered'))
    avg_confidence = sum(r.get('confidence', 0) for r in results) / len(results) if results else 0
    final_output = {
        'metadata': {
            'transcript_file': uploaded_file.name,
            'checklist_file': CHECKLIST_PATH,
            'total_items': len(results),
            'items_answered': items_answered,
            'items_not_found': len(results) - items_answered,
            'average_confidence': avg_confidence
        },
        'checklist': updated_checklist
    }
    excel_path = f"{run_dir}/{base_name}.xlsx"
    create_excel_report(final_output, excel_path)
    st.session_state['excel_path'] = excel_path
if 'excel_path' in st.session_state:
    with open(st.session_state['excel_path'], 'rb') as f:
        st.download_button(
            label="Download Excel Report",
            data=f.read(),
            file_name=os.path.basename(st.session_state['excel_path']),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )